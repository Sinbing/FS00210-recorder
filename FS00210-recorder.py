# code by: utf-8
# Sinbing code
import os
import sys
import serial
import datetime
import binascii
import openpyxl
from serial.tools import list_ports


def datetime_now():
    return datetime.datetime.now().strftime("%H:%M:%S")


def get_ports():
    ports = list_ports.comports()
    while True:
        p_name = ''
        for p in ports:
            p_name = p_name +p.device +' '
        if len(ports) > 1:
            print('搜索到串口设备： ' +p_name)
            user_input_ports = input('请输入需要使用的串口(如: COM1  输入"r"重新搜索)  \n: ')
            if user_input_ports != 'r':
                return user_input_ports

        else:
            print(f'监听唯一串口设备 {p.device}')
            return p.device


def excel_write(excel_path: str, excel_name: str, excel_row, data_PM100: str, data_PM25: str, data_PM10: str):
    # 尝试打开excel文件，若不存在则新建
    try:
        workbook = openpyxl.load_workbook(os.path.join(excel_path, excel_name))
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]

    # 写入Excel, 1行时创建表头
    if excel_row == 1:
        worksheet.cell(excel_row, 1, '记录时间')
        worksheet.cell(excel_row, 2, 'PM10 浓度(μg/m³)')
        worksheet.cell(excel_row, 3, 'PM2.5 浓度(μg/m³)')
        worksheet.cell(excel_row, 4, 'PM1.0 浓度(μg/m³)')
    worksheet.cell(excel_row+1, 1, datetime.datetime.now().strftime("%Y-%m-%d_%H:%M:%S"))
    worksheet.cell(excel_row+1, 2, data_PM100)
    worksheet.cell(excel_row+1, 3, data_PM25)
    worksheet.cell(excel_row+1, 4, data_PM10)

    # 保存Excel文件
    workbook.save(os.path.join(excel_path, excel_name))



def init():
    # 获取程序位置
    if hasattr(sys, "_MEIPASS"):
        progame_path = os.path.dirname(os.path.realpath(sys.executable))
    else:
        progame_path = os.path.split(os.path.realpath(__file__))[0]

    # 问问你要不要Excel输出
    while True:
        excel = input('是否需要输出Excel数据表\n 0: 不要\n 1: 要\n    输入数字: ')
        try:
            excel = int(excel)
        except:
            print('请只输入数字！\n')
        if excel == 0:
            excel_flag = False
            break
        elif excel == 1:
            excel_flag = True
            break
        else:
            print('    请重新输入')
    return progame_path, excel_flag



if __name__ == '__main__':
    # 初始化; 获取串口列表, 选择串口; 设置通信参数
    data_line, byte_count, excel_name, excel_row = '', 0, f'{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}_PM传感器数据.xlsx', 1
    progame_path, Excel_Output = init()
    ports = get_ports()
    ser = serial.Serial(ports, baudrate=9600, bytesize=8, timeout=2)
    # Main
    while True:
        data = ser.read()
        if data:
            hex_data = str(binascii.hexlify(data)).split('\'')[1]
            data_line += hex_data
            byte_count += 1
            # 根据数据特性 以64为单位分割数据(64字节一组/s)
            if byte_count == 64:
                #print(datetime_now() +' | '+ data_line)

                # 校验数据头，防止数据错位
                if data_line[:4] == '424d':
                    PM10 = int(data_line[23:24], 16)    # 数据04 PM1.0 大气环境下 μg/m³
                    PM25 = int(data_line[27:28], 16)    # 数据05 PM2.5 大气环境下 μg/m³
                    PM100 = int(data_line[31:32], 16)   # 数据06 PM10 大气环境下 μg/m³
                    print(f'{datetime_now()}: PM10: {PM100} μg/m³ | PM2.5: {PM25} μg/m³ | PM1.0: {PM10} μg/m³')
                    if Excel_Output:
                        excel_write(progame_path, excel_name, excel_row, PM100, PM25, PM10)
                        excel_row += 1
                # 次轮数据处理完毕，重置
                data_line, byte_count = '', 0